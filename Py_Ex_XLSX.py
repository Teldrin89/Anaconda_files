# different approach for handling the excel files - new format ".xlsx"
# import full module - openpyxl - in attempt to check if the intelli sense help would work (still not working,
# change back to importing sub-packages
from openpyxl import Workbook

# initialize new workbook - assign it to variable
wb = Workbook()
# activate 1st worksheet from initialized workbook
ws = wb.active
# create new worksheet and assign it to variable
ws1 = wb.create_sheet("New_Sheet")
# after name you could specify the place of new sheet - index
ws2 = wb.create_sheet("Newer_Sheet", 0)
# rename title of specified sheet
ws.title = "Next_Sheet"
# re-color sheet name tab
ws.sheet_properties.tabColor = "1072BA"
# save workbook - it will overwrite if there is already a workbook with that name
wb.save("testing.xlsx")
# initialize next workbook - new variable
wb2 = Workbook()
# create new worksheet in new workbook - remember to use proper variable
ws3 = wb2.create_sheet("First_Sheet", 0)
# save new workbook
wb2.save("new_doc.xlsx")

# TODO - opening of workbooks, I/O of data
