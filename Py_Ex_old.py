# 2 imported modules refer to work with excel spreadsheets from python - "xlwt" and "xlrd" - for older version of excel
# 1 imported module refer to work with new version of excel file (".xlsx") - "openpyxl"
# from datetime module import datetime method (function)
import xlwt
import xlrd
from datetime import datetime
import openpyxl

# creation of excel file
workbook = xlwt.Workbook(encoding="cp1250")
# creation of different sheets
sheet1 = workbook.add_sheet("Full_Report")
sheet2 = workbook.add_sheet("Expenses")
sheet3 = workbook.add_sheet("Income")
# filling cells in each sheet
sheet1.write(0, 0, "Some important report")
sheet2.write(1, 10, "We spent too much")
sheet3.write(0, 2, "But they payed us more")
sheet2.write(0, 0, "12.3")
sheet3.write(0, 0, "456")
# saving the file
workbook.save("report1.xls")
# opening the workbook
workbook = xlrd.open_workbook("report1.xls")
# printing out the cell values (from cells "A1") from each sheet
list_of_cells = []
for sheet in workbook.sheet_names():
    book = workbook.sheet_by_name(sheet)
    print(book.row_values(0)[0])
    list_of_cells.append(book.row_values(0)[0])

# checking out the type of variable read from cell
print(type(list_of_cells[1]))
# checking if variable read from cell is digit - NOT working with floating point numbers
print(list_of_cells[1].isdigit())

# customization of cells look - font type, size, attribute, color and format
style0 = xlwt.easyxf('font: name Times New Roman, color-index red, bold on, height 400', num_format_str='#,##0.00')
# create new style - "style1" - and assign to it a different type of formatting (date type)
style1 = xlwt.easyxf("", "DD-MM-YY")
# create new workbook with "xlwt" module
wb = xlwt.Workbook()
# create a new worksheet
ws = wb.add_sheet('A Test Sheet')
# assign several variables to certain cells - including formatting on some of them and function assign in others
ws.write(0, 0, "Test", style0)
ws.write(1, 0, datetime.now(), style1)
ws.write(2, 0, 1)
ws.write(2, 1, 1)
ws.write(2, 2, xlwt.Formula("A3+B3"))
# there is only one problem here - intellij seem to not find a help for "write" method from "xlwt" - no auto-fill
wb.save('example.xls')

# Exercise: prepare the excel file similar to the one described in .pdf file
wb2 = xlwt.Workbook()
ws2 = wb2.add_sheet("Hey")

for i in range(0, 5):
    for j in range(5):
        ws2.write(i, j, int(i))

ws2.write(6, 0, "Lukasz Cieslak")
style3 = xlwt.easyxf('font: underline on')
ws2.write(7, 0, "Python", style3)
wb2.save('numbers.xls')
